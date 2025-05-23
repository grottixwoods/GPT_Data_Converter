"""Модуль обработки и извлечения текста из документов."""
import os
import platform
import re
import shutil
import threading
from typing import Dict, List, Optional, Set, Tuple, Union

# Сторонние библиотеки
import openpyxl
import pandas as pd
import PyPDF2
import pytesseract
import textract
import xlrd
from bs4 import BeautifulSoup
from docx import Document
from docx.opc.exceptions import PackageNotFoundError
from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter
from pdf2image import convert_from_path
from pdfminer.pdfdocument import PDFDocument
from pdfminer.pdfparser import PDFParser, PSSyntaxError
from PIL import Image

if platform.system() == "Windows":
    from win32com.client import constants
    import win32com.client as win32

# Константы типов файлов
FILE_TYPES: Tuple[str, ...] = ('.docx', '.xlsx', '.ppt', '.odt')
IMAGE_FILES: Tuple[str, ...] = ('.jpg', '.png', '.jpeg', '.bmp', '.tif')

def handle_timeout() -> None:
    """Обработчик таймаута, вызывает исключение при превышении времени выполнения."""
    raise Exception("Превышено время ожидания операции")

def move_files_to_root(folder_path: str) -> None:
    """Перемещает все файлы из подпапок в корневую папку.
    
    Args:
        folder_path: Путь к корневой папке
    """
    for root, _, files in os.walk(folder_path):
        if root != folder_path:  # Исключаем корневую папку
            for file in files:
                src = os.path.join(root, file)
                dst = os.path.join(folder_path, file)
                if not os.path.exists(dst):
                    shutil.move(src, dst)

def convert_xls_to_xlsx(input_files: str) -> None:
    """Конвертирует файлы XLS в формат XLSX.
    
    Args:
        input_files: Директория с входными файлами
    """
    for filename in os.listdir(input_files):
        if filename.endswith('.xls'):
            xls_path = os.path.join(input_files, filename)
            wb = xlrd.open_workbook(xls_path)
            xlsx_path = os.path.join(input_files, filename[:-4] + '.xlsx')
            new_wb = Workbook()
            ws = new_wb.active

            for sheet_name in wb.sheet_names():
                sheet = wb.sheet_by_name(sheet_name)
                for row in range(sheet.nrows):
                    for col in range(sheet.ncols):
                        col_letter = get_column_letter(col + 1)
                        cell_value = sheet.cell_value(row, col)
                        ws[f"{col_letter}{row + 1}"] = cell_value

            new_wb.save(xlsx_path)
            os.remove(xls_path)

def convert_doc_to_docx(input_files: str) -> None:
    """Конвертирует файлы DOC в формат DOCX (только для Windows).
    
    Args:
        input_files: Директория с входными файлами
    """
    if platform.system() != "Windows":
        return

    for filename in os.listdir(input_files):
        if filename.endswith('.doc'):
            input_path = os.path.join(input_files, filename)
            word = win32.gencache.EnsureDispatch('Word.Application')
            doc = word.Documents.Open(input_path)
            doc.Activate()

            new_file_abs = os.path.abspath(input_path)
            new_file_abs = re.sub(r'\.\w+$', '.docx', new_file_abs)
            word.ActiveDocument.SaveAs(
                new_file_abs, FileFormat=constants.wdFormatXMLDocument
            )
            doc.Close(False)
            os.remove(input_path)

def extract_metadata(input_files: str, output_txt: str) -> None:
    """Извлекает метаданные из различных типов документов.
    
    Args:
        input_files: Директория с входными файлами
        output_txt: Директория для выходных файлов
    """
    df = pd.DataFrame(columns=['Meta', 'Path'])
    
    for filename in os.listdir(input_files):
        file_path = os.path.join(input_files, filename)
        metadata_dict: Dict[str, str] = {}
        
        if filename.endswith(".docx"):
            try:
                doc = Document(file_path)
                metadata = doc.core_properties
                metadata_dict = {
                    "Title": metadata.title,
                    "Author": metadata.author,
                    "Subject": metadata.subject,
                    "Keywords": metadata.keywords,
                    "Category": metadata.category,
                    "Comments": metadata.comments
                }
            except PackageNotFoundError:
                print("Ошибка: пакет не найден")
                
        elif filename.endswith(".pdf"):
            with open(file_path, 'rb') as pdf_file:
                parser = PDFParser(pdf_file)
                try:
                    doc = PDFDocument(parser)
                    metadata_dict = doc.info[0]
                except (PSSyntaxError, AttributeError):
                    metadata_dict = {}
                    
        elif filename.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
            try:
                wb = openpyxl.load_workbook(file_path)
                metadata_dict = {
                    'authors': wb.properties.creator,
                    'theme': wb.properties.title,
                    'tags': wb.properties.keywords,
                    'category': wb.properties.category,
                    'comments': wb.properties.description
                }
            except TypeError:
                print("Ошибка при загрузке Excel файла")
                
        df = df._append({
            'Meta': metadata_dict,
            'Path': os.path.join(output_txt, f'{filename[:-5]}.txt')
        }, ignore_index=True)
        
    df.to_csv('dataframe.csv', index=False)

def extract_text_from_documents(input_files: str, output_txt: str) -> None:
    """Извлекает текст из различных типов документов.
    
    Args:
        input_files: Директория с входными файлами
        output_txt: Директория для выходных файлов
    """
    for filename in os.listdir(input_files):
        print(f'Обработка файла: {filename}')
        file_path = os.path.join(input_files, filename)
        
        if filename.endswith('.pdf'):
            process_pdf_file(filename, input_files, output_txt)
        elif filename.endswith(FILE_TYPES):
            process_textract_file(filename, input_files, output_txt)
        elif filename.endswith('.html'):
            process_html_file(filename, input_files, output_txt)
        elif filename.endswith(IMAGE_FILES):
            process_image_file(filename, input_files, output_txt)
            
        os.remove(file_path)

def process_pdf_file(filename: str, input_files: str, output_txt: str) -> None:
    """Обрабатывает PDF файлы с использованием PyPDF2 или OCR.
    
    Args:
        filename: Имя PDF файла
        input_files: Входная директория
        output_txt: Выходная директория
    """
    with open(os.path.join(input_files, filename), 'rb') as file:
        has_text = False
        try:
            pdf_reader = PyPDF2.PdfReader(file)
            for page in pdf_reader.pages:
                if page.extract_text():
                    has_text = True
                    break
        except Exception:
            has_text = None
            
    if not has_text:
        process_pdf_with_ocr(filename, input_files, output_txt)
    else:
        process_pdf_with_textract(filename, input_files, output_txt)

def process_pdf_with_ocr(filename: str, input_files: str, output_txt: str) -> None:
    """Обрабатывает PDF с использованием OCR, когда извлечение текста не удалось.
    
    Args:
        filename: Имя PDF файла
        input_files: Входная директория
        output_txt: Выходная директория
    """
    pages = convert_from_path(os.path.join(input_files, filename), 400)
    text = ""
    try:
        timer = threading.Timer(600, handle_timeout)
        timer.start()
        for img_blob in pages:
            text += pytesseract.image_to_string(img_blob, lang='rus') + '\n'
        timer.cancel()
    except Exception as e:
        print(f"Ошибка OCR: {e}")
    finally:
        save_text_to_file(filename, text, output_txt)

def process_pdf_with_textract(filename: str, input_files: str, output_txt: str) -> None:
    """Обрабатывает PDF с использованием textract.
    
    Args:
        filename: Имя PDF файла
        input_files: Входная директория
        output_txt: Выходная директория
    """
    try:
        text = textract.process(os.path.join(input_files, filename)).decode('utf-8')
    except Exception:
        print('Ошибка при обработке textract')
        text = 'Ошибка извлечения текста'
    save_text_to_file(filename, text, output_txt)

def process_textract_file(filename: str, input_files: str, output_txt: str) -> None:
    """Обрабатывает файлы с использованием textract.
    
    Args:
        filename: Имя файла
        input_files: Входная директория
        output_txt: Выходная директория
    """
    try:
        text = textract.process(os.path.join(input_files, filename)).decode('utf-8')
    except Exception:
        print('Ошибка при обработке textract')
        text = 'Ошибка извлечения текста'
    save_text_to_file(filename, text, output_txt)

def process_html_file(filename: str, input_files: str, output_txt: str) -> None:
    """Обрабатывает HTML файлы.
    
    Args:
        filename: Имя HTML файла
        input_files: Входная директория
        output_txt: Выходная директория
    """
    file_path = os.path.join(input_files, filename)
    with open(file_path, 'rb') as html_file:
        html_content = html_file.read()
        soup = BeautifulSoup(html_content, 'html.parser')
        text = soup.get_text()
        save_text_to_file(filename, text, output_txt)

def process_image_file(filename: str, input_files: str, output_txt: str) -> None:
    """Обрабатывает изображения с использованием OCR.
    
    Args:
        filename: Имя файла изображения
        input_files: Входная директория
        output_txt: Выходная директория
    """
    image = Image.open(os.path.join(input_files, filename))
    text = pytesseract.image_to_string(image, lang='rus')
    save_text_to_file(filename, text, output_txt)

def save_text_to_file(filename: str, text: str, output_txt: str) -> None:
    """Сохраняет извлеченный текст в файл.
    
    Args:
        filename: Имя исходного файла
        text: Извлеченный текст
        output_txt: Выходная директория
    """
    new_filename = os.path.splitext(filename)[0] + '.txt'
    with open(os.path.join(output_txt, new_filename), 'w', encoding='utf-8') as f:
        f.write(text)

def clean_text_files(output_txt: str) -> None:
    """Очищает и форматирует текстовые файлы.
    
    Args:
        output_txt: Директория с текстовыми файлами
    """
    for filename in os.listdir(output_txt):
        if filename.endswith('.txt'):
            filepath = os.path.join(output_txt, filename)
            
            # Удаление пустых строк
            with open(filepath, 'r', encoding='utf-8') as f:
                lines = [line for line in f.readlines() if line.strip()]
            with open(filepath, 'w', encoding='utf-8') as f:
                f.writelines(lines)
            
            # Очистка форматирования текста
            with open(filepath, 'r', encoding='utf-8') as file:
                text = file.read()
            
            # Применение правил очистки текста
            text = re.sub(r'([а-я])([А-Я])', r'\1 \2', text)  # Добавление пробела между русскими словами
            text = re.sub(r'(\w+)-\n(\w+)', r'\1\2', text)    # Исправление переносов слов
            text = re.sub(r'(?<=\w)– ', '', text)             # Удаление длинных тире
            text = re.sub(r'(?<=\w)- ', '', text)             # Удаление дефисов
            text = text.replace(" ", " ")                      # Нормализация пробелов
            text = text.replace("", "")                        # Удаление невидимых символов
            text = text.replace("", " ")                       # Замена специальных символов
            text = re.sub(r'(?<=\b\w)\s(?=\w\b)', '', text)   # Удаление лишних пробелов
            
            with open(filepath, 'w', encoding='utf-8') as file:
                file.write(text)            